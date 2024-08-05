from datetime import datetime
from openpyxl import load_workbook as load_wb
import win32com.client as win32
import customtkinter as ctk
from tkinter import filedialog, messagebox
import sqlite3 as sql

# -------------------------------------
tabela = sql.connect('database.db')
cursor = tabela.cursor()


# -------------------------------------

def inserir(x):
    cursor.execute('SELECT * FROM dados')
    registros1 = cursor.fetchall()
    if registros1:
        cursor.execute("DELETE FROM dados")
    cursor.execute("INSERT INTO dados VALUES (?)", (x,))
    tabela.commit()


# -=-=-=-=--=--=--=--=--=--=--=--=--=--=-=--=-=-Janela--=--=--=--=--=--=--=--=--=--=--=--=--=--=--=--=--=-

def abrir_janela():
    global data_entry, diretorio_entry, janela
    janela_principal.destroy()

    # Janelas
    janela = ctk.CTk()
    janela.title("Receber Dados do Usuário")
    janela.geometry("400x300")
    janela.maxsize(width=400, height=200)
    janela.minsize(width=400, height=200)

    # entrada do diretorio
    diretorio_label = ctk.CTkLabel(janela, text="Diretório:")
    diretorio_label.place(x=10, y=60)
    diretorio_frame = ctk.CTkFrame(janela)
    diretorio_frame.place(x=150, y=60)

    diretorio_entry = ctk.CTkEntry(diretorio_frame)
    diretorio_entry.grid(row=0, column=0, sticky="ew")
    diretorio_button = ctk.CTkButton(diretorio_frame, text="Selecionar", command=selecionar_diretorio, width=50)
    diretorio_button.grid(row=0, column=1, padx=5, sticky="w")

    # botão de capturar dados
    capturar_button = ctk.CTkButton(janela, text="Capturar Dados", command=capturar_dados)
    capturar_button.place(x=150, y=130)

    janela.mainloop()


def selecionar_diretorio():
    global diretorio1
    diretorio1 = str(filedialog.askopenfilename(
        title="Selecione o arquivo",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls")]))
    if not diretorio1:
        messagebox.showerror("Erro", "Selecione o Diretorio ou Feche a aplicação,")
    else:
        diretorio_entry.delete(0, ctk.END)
        diretorio_entry.insert(0, diretorio1)


def capturar_dados():
    inserir(diretorio1)
    janela.destroy()


# -=-=-=-=--=--=--=--=--=--=--=--=--=--=-=--=--Importar--=--=--=--=--=--=--=--=--=--=--=--=--=--=--=--=--=

def obter_numero_de_emails(data):
    outlook = win32.Dispatch('Outlook.Application').GetNamespace('MAPI')
    inbox = outlook.GetDefaultFolder(6)  # Ele vai pegar a caixa n° 6 do outlook, no caso o inbox

    lista = data_entry.get().split('/')
    print(lista)
    dia, mes, ano = lista

    dia = int(dia)
    mes = int(mes)
    ano = int(ano)
    data_inicial = datetime(ano, mes, dia)
    data_final = datetime(ano, mes, dia + 1)

    filtro_inicial = f"[ReceivedTime] >= '{data_inicial.strftime('%d/%m/%Y 00:00')}'"
    filtro_final = f"[ReceivedTime] < '{data_final.strftime('%d/%m/%Y 00:00')}'"
    filtro_total = f"({filtro_inicial}) AND ({filtro_final})"

    emails = inbox.Items.Restrict(filtro_total)
    return emails.Count


# -=-=-=--=--=--=--=--=--=--=--=--=--=--=--=--Exportar--=--=--=--=--=--=--=--=--=--=--=--=--=--=--=--=--=-

def criar_p():
    global data
    data = data_entry.get()  # Pega a data da caixinha
    if not data:
        messagebox.showerror("Erro", "A data deve ser preenchida.")
        return
    try:
        # verificação
        datetime.strptime(data, '%d/%m/%Y')
    except ValueError:
        messagebox.showerror("Erro de Formato", "A data deve estar no formato dd/mm/yyyy.")
        return

    n_mensagens = obter_numero_de_emails(data)  # obtem os numeros com a data

    cursor.execute("SELECT * FROM dados")
    registro1 = cursor.fetchall()
    diretorio = registro1[0][0]
    wb = load_wb(diretorio)
    ws = wb.worksheets[0]
    if ws['A2']:
        ws.insert_rows(2)
    if ws['A1'].value is None:
        ws['A1'] = 'Data'
    if ws['B1'].value is None:
        ws['B1'] = 'Emails'
    ws['A2'] = f'{data}:'
    ws['B2'] = str(n_mensagens)
    wb.save(diretorio)
    janela_principal.destroy()


# janela principal
janela_principal = ctk.CTk()
janela_principal.title("Aplicação Principal")
janela_principal.geometry("200x250")
janela_principal.maxsize(width=200, height=250)
janela_principal.minsize(width=200, height=250)

# data labels
data_label = ctk.CTkLabel(janela_principal, text="Data (dd/mm/yyyy):")
data_label.pack(pady=10)
data_entry = ctk.CTkEntry(janela_principal)
data_entry.pack(pady=5)

# botão de atualizar
atualizar_button = ctk.CTkButton(janela_principal, text="Atualizar Dados", command=abrir_janela)
atualizar_button.pack(pady=15)

# botão de continuar
continuar_b = ctk.CTkButton(janela_principal, text="Criar Planilha", command=criar_p)
continuar_b.pack(pady=20)

janela_principal.mainloop()
